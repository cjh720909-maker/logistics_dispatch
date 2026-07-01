from openpyxl import load_workbook

from db_local import local_engine


def cell_value(ws, row, col):
    value = ws.cell(row=row, column=col).value

    if value is None:
        return ""

    return str(value).strip()


def read_daesang_excel(file, preview_limit=None):
    wb = load_workbook(file, data_only=True)

    if "대상원본" not in wb.sheetnames:
        raise Exception("대상원본 시트를 찾을 수 없습니다.")

    ws = wb["대상원본"]

    rows = []

    for i in range(11, ws.max_row + 1):
        delivery_date = cell_value(ws, i, 2)

        if not delivery_date:
            break

        row = {
            "delivery_date": delivery_date,
            "delivery_name": cell_value(ws, i, 4),
            "delivery_code": cell_value(ws, i, 5),
            "seller_name": cell_value(ws, i, 6),
            "delivery_type": cell_value(ws, i, 7),
            "product_code": cell_value(ws, i, 8),
            "product_name": cell_value(ws, i, 9),
            "unit_price": cell_value(ws, i, 10),
            "vat_type": cell_value(ws, i, 11),
            "weight": cell_value(ws, i, 12),
            "ipsu": cell_value(ws, i, 13),
            "qty": cell_value(ws, i, 14),
            "weight_unit": cell_value(ws, i, 15),
            "nap_no": cell_value(ws, i, 16),
            "order_no": cell_value(ws, i, 17),
        }

        rows.append(row)

        if preview_limit and len(rows) >= preview_limit:
            break

    return rows


def preview_daesang_excel(file):
    rows = read_daesang_excel(file, preview_limit=None)
    validation = validate_daesang_rows(rows)

    return {
        "rows": rows[:20],
        "all_rows": rows,
        "total_count": len(rows),
        "validation": validation,
    }

def validate_daesang_rows(rows):
    delivery_codes = sorted({
        row["delivery_code"]
        for row in rows
        if row.get("delivery_code")
    })

    product_codes = sorted({
        row["product_code"]
        for row in rows
        if row.get("product_code")
    })

    delivery_map = {}
    product_set = set()

    with local_engine.connect() as conn:
        if delivery_codes:
            placeholders = ",".join(["?"] * len(delivery_codes))

            result = conn.exec_driver_sql(
                f"""
                SELECT
                    code,
                    name
                FROM delivery
                WHERE code IN ({placeholders})
                """,
                tuple(delivery_codes)
            )

            for row in result.mappings().all():
                delivery_map[row["code"]] = row["name"]

        if product_codes:
            placeholders = ",".join(["?"] * len(product_codes))

            result = conn.exec_driver_sql(
                f"""
                SELECT
                    code
                FROM product
                WHERE code IN ({placeholders})
                """,
                tuple(product_codes)
            )

            product_set = {
                row["code"]
                for row in result.mappings().all()
            }

    new_deliveries = {}
    name_mismatches = {}
    new_products = {}

    for row in rows:
        delivery_code = row.get("delivery_code")
        delivery_name = row.get("delivery_name")
        product_code = row.get("product_code")
        product_name = row.get("product_name")

        if delivery_code:
            master_name = delivery_map.get(delivery_code)

            if not master_name:
                new_deliveries[delivery_code] = delivery_name

            elif master_name != delivery_name:
                name_mismatches[delivery_code] = {
                    "excel_name": delivery_name,
                    "master_name": master_name,
                }

        if product_code and product_code not in product_set:
            new_products[product_code] = product_name

    return {
        "new_deliveries": new_deliveries,
        "name_mismatches": name_mismatches,
        "new_products": new_products,
        "can_save": (
            len(new_deliveries) == 0
            and len(name_mismatches) == 0
            and len(new_products) == 0
        )
    }

def match_base_category(rule_value, row_base_category):
    if not rule_value or rule_value.strip() == "*":
        return True

    allowed = [
        item.strip()
        for item in rule_value.split(",")
        if item.strip()
    ]

    return row_base_category in allowed